# reading excel file using openpyxl into python dictionary
# Date: 2022/01/10
import re
import pprint
import openpyxl

# function to format dictionary key
def fmt_key(key_name):
    # remove leading and trailling spaces
    key_name = key_name.strip()
    # replace spaces with underscores
    key_name = key_name.lower().replace(" ", "_")
    # remove non alphanumeric characters
    key_name = re.sub(r"[^a-zA-Z0-9_]", "", key_name)
    # return value
    return key_name


# set location of file
loc = "movielist_samplefile.xlsx"

# open workbook, read only property allows for more consisttent and faster read times esecially when working on larger excel files
wb = openpyxl.load_workbook(loc, read_only=True)

# print sheetname
print(wb.sheetnames)

# set active sheet
sheet = wb.active

# get list of column names
key_names = []
for x in range(1, sheet.max_column + 1):
    key_names.append(fmt_key(sheet.cell(1, x).value))

# convert rows to dictionary
row_values = []
# iterate over rows
for row in sheet.iter_rows(min_row=2, values_only=True):
    res = []
    # set dictinary key:value for each column
    res = {key_names[i]: row[i] for i in range(len(key_names))}
    # append dictinary to list
    row_values.append(res)

# print values
pprint.pprint(row_values)
